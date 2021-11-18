########## Angkan Baidya ##########
########## abaidya ##############
########## 112309655 #############

import openpyxl
from openpyxl.utils.exceptions import InvalidFileException


import os
import re


class ExcelFileManager:
    def __init__(self, filename):
        if os.path.isdir(filename) == True:
            raise ValueError("file is a directory")
        self.filename = filename

    def write_sheet(self, data, sheetname, wb=None, save=True):
        if save == True:
            myworkbook = openpyxl.workbook.Workbook()
        else:
            myworkbook = wb
        print("inwrite sheet" ,data)
        print("inwrite sheet",sheetname)
        listofnumbers = []
        listofnames = []
        new_list = []
        hitastring = False
        if sheetname in myworkbook.sheetnames:
            ws1 = myworkbook[sheetname]
        else:
            ws1 = myworkbook.create_sheet(sheetname)
        print("total worksheets",myworkbook.sheetnames)
        counter = 0
        nextiter=False
        for i in data:
            nextiter = False
            for z in data:
                if isinstance(z[1],str):
                    listofnames.append(z[0])
                    try:
                        numbertoadd = float(z[1])
                    except ValueError:
                        return -1
                    listofnumbers.append(numbertoadd)
                    hitastring = True
                if (isinstance(z[1],int) | isinstance(z[1],float)):
                    listofnames.append(z[0])
                    listofnumbers.append(z[1])
            if len(listofnumbers) > 0 and hitastring == True:
                new_list = [(listofnames[i], listofnumbers[i]) for i in range(0, len(listofnames))]
                tuple(new_list)
                break
            
            for cellz in ws1['A']:
                if cellz.value == None:
                    ws1.cell(row = 1, column =1).value = i[0]
                    counter = counter +1
                    nextiter= True
                    continue
            for cell in ws1['A']:
                if cell.value== i[0]:
                    for x in ws1.iter_rows(min_row = cell.row,max_row = cell.row,min_col = 1,max_col = 10):
                        for cels in x:
                            if cels.value is None:
                                cels.value = i[1]
                                nextiter = True
                                break
            if (nextiter == False):
                ws1.append(i)
                counter = counter + 1
        myworkbook.save(self.filename)
        newcounter = 0
        nextiter2 = False
        if(len(new_list)> 1):
            for i in new_list:
                nextiter2 = False
                for cellz in ws1['A']:
                    if cellz.value == None:
                        ws1.cell(row = 1, column =1).value = i[0]
                        counter = counter +1
                        nextiter= True
                        continue
                for cell in ws1['A']:
                    print("CellValue:",cell.value)
                    if cell.value== i[0]:
                        print(cell.row)
                        for x in ws1.iter_rows(min_row = cell.row,max_row = cell.row,min_col = 1,max_col = 10):
                            for cels in x:
                                print(cels.value)
                                if cels.value is None:
                                    cels.value = i[1]
                                    nextiter2 = True
                                    break
                if (nextiter2 == False):
                    ws1.append(i)
                    newcounter = newcounter + 1
                myworkbook.save(self.filename)
                for cell in ws1['A']:
                    print(cell.value)
                    print(cell.column,cell.row)
                for cell in ws1['1']:
                    print(cell.value)
                    print(cell.column,cell.row)
        myworkbook.save(self.filename)
        return 0






            

    def write_sheets(self, data):
        listofname = []
        listofnumbers = []
        listofsheetnames = []
        new_list = []
        counter = 0
        myworkbook = None
        for i in data:
            listofname.clear()
            listofnumbers.clear()
            listofsheetnames.clear()
            listofname.append(i[0])
            try:
                numbertoadds = float(i[1])
            except ValueError:
                 return -1
        for i in data:
            listofname.clear()
            listofnumbers.clear()
            listofsheetnames.clear()
            listofname.append(i[0])
            try:
                numbertoadds = float(i[1])
            except ValueError:
                myworkbook.save(self.filename)
                return -1
            listofnumbers.append(numbertoadds)
            listofsheetnames.append(i[2])
            new_list = [(listofname[i], listofnumbers[i]) for i in range(0, len(listofname))]
            tuple(new_list)
            print(new_list)
            print(listofsheetnames)
            sheetname = listofsheetnames[0]
            if counter == 0:
                 myworkbook = openpyxl.workbook.Workbook()
                 self.write_sheet(new_list,sheetname,myworkbook,False)
            else:
                self.write_sheet(new_list,sheetname,myworkbook,False)
                myworkbook.save(self.filename)
            counter = counter +1
        

            

    def get_sheet_ave(self,sheetname):
        try:
            wb = openpyxl.load_workbook(self.filename)
            sheet = wb[sheetname]
        except FileNotFoundError:
            return []
        except KeyError:
            return []
        except ValueError:
            return []
        except InvalidFileException:
            return []

        listofnames = []
        listofnames2 = []
        listofnumbers = []
        counter = 0
        for i in sheet['A']:
            listofnames2.append(i.value)


        for i in sheet['A']:
            listofnames.append(i.value)
            for x in sheet.iter_rows(min_row = i.row,max_row = i.row,min_col = 1,max_col = 20):
                numbercounter = 0
                counterforavg = 0
                for cels in x:
                    if cels.value is None:
                        break
                    if isinstance(cels.value,str):
                        continue
                    else:
                        numbercounter = numbercounter + cels.value
                        counterforavg = counterforavg + 1
                numbercounter = numbercounter/counterforavg
                listofnumbers.append(numbercounter)
        new_list = [(listofnames[i], listofnumbers[i]) for i in range(0, len(listofnames))]
        tuple(new_list) 
        print(new_list)
        return new_list


    def get_workbook_ave(self, pattern=None):
        listofpeople = []
        listofavg = []
        wb = openpyxl.load_workbook(self.filename)
        print("THESE ARE THE SHEETNAMES",wb.sheetnames)
        pattern = "^(C|I)SE337$"
        meanofmeans = 0
        meanofmeanslist = []
        meanofmeansperson = []
        for i in wb.sheetnames:
            print(i)
            word = i 
            result =re.match(pattern,word)
            print(result)
            if result:
                sample = self.get_sheet_ave(i)
                print(sample)
                alist = list(sample)
                for i in alist:
                    listofpeople.append(i[0])
                    listofavg.append(i[1])
        for i in listofavg:
            float(i)
            meanofmeans = meanofmeans + i 
            meanofmeans = meanofmeans / len(listofavg)
            meanofmeanslist.append(meanofmeans)
        for i in listofpeople:
            check = i.lower()
            if len(meanofmeansperson) == 0:
                meanofmeansperson.append(i)
            for i in listofpeople:
                if i not in meanofmeansperson:
                    meanofmeansperson.append(i)
        new_list = [(meanofmeansperson[i], meanofmeanslist[i]) for i in range(0, len(meanofmeansperson))]
        tuple(new_list)
        print(new_list)
        return(new_list)



        

